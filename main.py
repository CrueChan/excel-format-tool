import openpyxl
from openpyxl.styles import PatternFill, Protection, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

def format_workbook(file_path):
    """
    处理Excel工作簿的格式
    
    Args:
        file_path (str): Excel文件路径
    """
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)
    
    # 定义填充样式
    header_gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # 深灰色
    content_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 浅灰色
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # 黄色
    
    # 处理每个工作表
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"处理工作表: {sheet}")
        
        # 找到分割列（包含"是否使用（必填）"的列）
        split_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and isinstance(cell_value, str) and '是否使用（必填）' in cell_value.replace('\n', ''):
                split_col = col
                break
        
        if split_col is None:
            print(f"Warning: 在工作表 {sheet} 中未找到包含'是否使用（必填）'的列")
            continue
        
        # 创建数据验证规则
        dv = DataValidation(
            type='list',
            formula1='"使用,禁用（注销）,已调离本单位"',  # 直接使用字符串列表
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
            promptTitle='使用状态选择',
            prompt='请从下拉列表中选择一个选项',
            errorTitle='输入错误',
            error='输入的值无效，请使用下拉列表选择一个有效的选项。'
        )
        # 强制设置显示属性
        dv.validation_type = 'list'
        dv._prompt_title = '使用状态选择'
        dv._prompt = '请从下拉列表中选择一个选项'
        dv._error_title = '输入错误'
        dv._error = '输入的值无效，请使用下拉列表选择一个有效的选项。'
        dv.errorStyle = 'stop'
        
        # 添加数据验证到工作表
        ws.add_data_validation(dv)
        
        # 为该列的所有数据行添加数据验证
        status_column_letter = get_column_letter(split_col)
        dv.add(f'{status_column_letter}2:{status_column_letter}{ws.max_row}')
        
        # 锁定工作表
        ws.protection.sheet = True
        ws.protection.password = 'E5T647kc'  # 设置保护密码
        
        # 设置所有单元格的对齐方式和自动换行
        wrap_alignment = Alignment(wrap_text=True, vertical='center')
        
        # 获取每列的最大字符数
        max_lengths = {}
        for col in range(1, ws.max_column + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    # 考虑换行符
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        if len(line) > max_length:
                            max_length = len(line)
            max_lengths[col] = max_length
        
        # 处理每一行
        for row in range(1, ws.max_row + 1):
            # 重置行高（使用默认值）
            if row in ws.row_dimensions:
                del ws.row_dimensions[row]
                
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                # 设置对齐方式和自动换行
                cell.alignment = wrap_alignment
                
                if col < split_col:  # 分割列之前的列
                    if row == 1:  # 标题行
                        cell.fill = header_gray_fill
                    else:  # 内容行
                        cell.fill = content_gray_fill
                    # 锁定单元格
                    cell.protection = Protection(locked=True)
                elif row == 1:  # 分割列及之后的标题行
                    cell.fill = yellow_fill
                    cell.protection = Protection(locked=True)
                else:  # 分割列及之后的内容
                    cell.protection = Protection(locked=False)
        
        # 自动调整列宽
        for col in range(1, ws.max_column + 1):
            column = get_column_letter(col)
            # 序号列（第一列）使用自动宽度
            if col == 1:
                adjusted_width = max_lengths[col] + 2  # 序号列使用精确宽度
            else:
                # 其他列限制最大宽度为50个字符
                adjusted_width = min(max_lengths[col] + 2, 50)
                # 确保最小宽度不小于8个字符
                adjusted_width = max(adjusted_width, 8)
            ws.column_dimensions[column].width = adjusted_width
        
        # 设置为普通视图模式
        ws.sheet_view.view = 'normal'
    
    # 保存修改
    wb.save(file_path)

def process_all_files(folder_path):
    """
    处理文件夹中的所有Excel文件
    
    Args:
        folder_path (str): 包含Excel文件的文件夹路径
    """
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            print(f"处理文件: {filename}")
            try:
                format_workbook(file_path)
                print(f"成功处理文件: {filename}")
            except Exception as e:
                print(f"处理文件 {filename} 时出错: {str(e)}")

if __name__ == '__main__':
    # 使用当前目录
    folder_path = '按部门拆分'  # 当前目录
    process_all_files(folder_path)